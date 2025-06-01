from openpyxl import load_workbook
from os.path import exists
from time import sleep

esc = 0
deadtime = 1


def listar_fornecedores():
    posi = 1
    for sheets in wb.sheetnames:
        sheet = wb[sheets]
        print(f'{posi} - {sheet.title}')
        posi += 1


def adicionaritem(item, qty, prazo, position):
    novoitem = item
    newqty = qty
    newprazo = prazo
    date = f"{newprazo[:2]}/{newprazo[2:4]}/{newprazo[4:]}"

    sheet[f'B{position}'] = novoitem
    sheet[f'C{position}'] = newqty
    sheet[f'F{position}'] = date


def atualizaritem(position):
    print(f"|   {sheet[f'B{cont}'].value} = {sheet[f'C{cont}'].value} | "
          f"Prazo: {sheet[f'F{cont}'].value}"
          )
    newqty = str(input("Qty: "))
    newprazo = str(input("Prazo ( DDMMAAAA ): "))
    date = f"{newprazo[:2]}/{newprazo[2:4]}/{newprazo[4:]}"
    if newqty == "":
        pass
    else:
        sheet[f'C{cont}'] = newqty
    if newprazo == "":
        pass
    else:
        sheet[f'F{cont}'] = date


def delete_cell_shift_up(ws, start_row, col_start='B', col_end='F'):
    max_row = ws.max_row
    col_range = range(ord(col_start), ord(col_end) + 1)

    for row in range(start_row, max_row):
        for col in col_range:
            current_cell = f'{chr(col)}{row}'
            next_cell = f'{chr(col)}{row + 1}'
            ws[current_cell].value = ws[next_cell].value
    for col in col_range:
        ws[f'{chr(col)}{max_row}'].value = None
    wb.save('material_externo.xlsx')


print('=-' * 20)
print('             APP PRAZOS')
print('=-' * 20)
while True:
    try:
        wb = load_workbook('material_externo.xlsx')
        break
    except FileNotFoundError:
        print("Arquivo ( material_externo.xlsx ) não encontrado!")

'''
while True:
    planilha = str(input('Digite o nome do seu arquivo excel ( xlsx ): '))
    if exists(planilha):
        wb = load_workbook(planilha)
        print('\033[32mArquivo carregado com sucesso.\033[0m')
        print('=-' * 20)
        break
    else:
        print("\033[31mArquivo NÃO encontrado.\033[0m")
'''

while esc not in ['1', '2', '3']:
    print('-' * 40)
    print('''
    1 - Atualizar prazos
    2 - Listar Itens
    3 - Adicionar Item
    4 - Deletar item
    5 - Sair
    ''')
    esc = str(input('Escolha: '))
    while True:
        if esc == '1':
            cont = 4
            for sheets in wb.sheetnames:
                sheet = wb[sheets]
                print("\n")
                print(f'{sheet.title}:')
                while True:
                    if sheet[f'B{cont}'].value is None:
                        if sheet[f'B4'].value is None:
                            add = 0
                            pos = 4
                            while add != 'n':
                                add = str(input(f"Deseja adicionar algum item a {sheet.title} (S/N)? ")).lower()
                                if add == 's':
                                    nitem = str(input("Item: "))
                                    nqty = str(input("Qty: "))
                                    nprazo = str(input("Prazo ( DDMMAAAA ): "))
                                    adicionaritem(nitem, nqty, nprazo, pos)
                                pos += 1
                        break
                    else:
                        atualizaritem(cont)

                    cont += 1
                cont = 4
            wb.save('material_externo.xlsx')
            sleep(deadtime)

        if esc == '2':
            cont = 4
            for sheets in wb.sheetnames:
                sheet = wb[sheets]
                print("\n")
                print(f'{sheet.title}:')
                while True:
                    if sheet[f'B{cont}'].value is None:
                        break
                    else:
                        print(f"|   {sheet[f'B{cont}'].value} = {sheet[f'C{cont}'].value} | "
                              f"Prazo: {sheet[f'F{cont}'].value}"
                              )
                    cont += 1
                cont = 4
            sleep(deadtime)

        if esc == '3':
            posicao = 1
            cont = 4
            listar_fornecedores()
            suplier = int(input("Qual fornecedor você deseja adiciona um item: ")) - 1
            while True:
                sheet = wb.worksheets[suplier]
                print(sheet[f'B{cont}'].value)
                if sheet[f'B{cont+1}'].value is None:
                    nitem = str(input("Item: "))
                    nqty = str(input("Qty: "))
                    nprazo = str(input("Prazo ( DDMMAAAA ): "))
                    adicionaritem(nitem, nqty, nprazo, cont+1)
                    sair = str(input("Deseja adicionar mais algum item (S/N): ")).lower()
                    if sair == 'n':
                        break
                    else:
                        cont += 1
                else:
                    cont += 1
        wb.save('material_externo.xlsx')
        sleep(deadtime)

        if esc == '4':
            cont = 4
            cont2 = 4
            posicao = 1
            listar_fornecedores()
            posicao2 = 1
            suplier = int(input("Qual fornecedor você deseja deletar: "))-1
            while True:
                while True:
                    sheet2 = wb.worksheets[suplier]
                    if sheet2[f'B{cont2}'].value is None:
                        break
                    else:
                        print(f"{posicao2} - {sheet2[f'B{cont2}'].value}")
                        cont2 += 1
                        posicao2 += 1
                print("Digite 0 para cancelar")
                deletar = int(input("Qual deseja deletar: ")) + 3
                if deletar - 3 == 0:
                    break
                delete_cell_shift_up(sheet2, deletar)
                if str(input("Deseja deletar outro item (S/N): ")).lower() == 'n':
                    break

        if esc == '5':
            print('=-' * 20)
            print("By: Alexandre Fagundes | 2025")
            print('=-' * 20)
            exit()
        esc = 0
        break


''' material_externo.xlsx '''