from openpyxl import load_workbook

planilha = load_workbook('material_externo.xlsx')
print('\n')
print(planilha.sheetnames)
print('\n')